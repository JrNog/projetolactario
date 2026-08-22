import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import re
import os

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

def parse_js_array(filepath, var_name):
    with open(filepath, 'r', encoding='utf-8') as f:
        code = f.read()
    
    # Encontra "const VAR_NAME = [" até o fechamento "]"
    match = re.search(r'const\s+' + var_name + r'\s*=\s*(\[[\s\S]*?\])(?:\.map|;|\n\s*if)', code)
    if not match:
        return []
    
    raw = match.group(1)
    # Remove comentários JS
    raw_clean = re.sub(r'//.*', '', raw)
    
    # 1. Tenta parsear direto se já estiver em formato JSON válido
    try:
        clean_strict = re.sub(r',\s*([\]}])', r'\1', raw_clean)
        return json.loads(clean_strict)
    except Exception:
        pass

    # 2. Se falhar, formata chaves não citadas
    try:
        quoted = re.sub(r'([{\s,])([a-zA-Z_][a-zA-Z0-9_]*)\s*:', r'\1"\2":', raw_clean)
        clean_quoted = re.sub(r',\s*([\]}])', r'\1', quoted)
        return json.loads(clean_quoted)
    except Exception as e:
        print(f"Erro ao parsear {var_name}:", e)
        return []

def create_database_spreadsheet():
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Arial", size=10)
    
    # -------------------------------------------------------------
    # 1. ABA DB_Censo (Censo / Relação de Pacientes)
    # -------------------------------------------------------------
    ws_censo = wb.create_sheet(title="DB_Censo")
    headers_censo = [
        "id", "rh", "nome", "enfermaria", "enfermariaNome", "leito",
        "dietaId", "dietaNome", "volumeMl", "vezesDia", "via",
        "dispositivo", "espessanteObs", "calCalorico", "horarioInicio",
        "suspenso", "updatedAt"
    ]
    ws_censo.append(headers_censo)
    
    header_fill_censo = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    for col_idx in range(1, len(headers_censo) + 1):
        cell = ws_censo.cell(row=1, column=col_idx)
        cell.fill = header_fill_censo
        cell.font = header_font
        cell.alignment = align_center

    mock_censo_path = os.path.join(BASE_DIR, "js", "data", "mock-censo.js")
    censo_data = parse_js_array(mock_censo_path, "MOCK_CENSO")
    for item in censo_data:
        row = [
            item.get("id", ""),
            item.get("rh", ""),
            item.get("nome", ""),
            item.get("enfermaria", ""),
            item.get("enfermariaNome", ""),
            item.get("leito", ""),
            item.get("dietaId", ""),
            item.get("dietaNome", ""),
            item.get("volumeMl", 0),
            item.get("vezesDia", 0),
            item.get("via", ""),
            item.get("dispositivo", ""),
            item.get("espessanteObs", ""),
            item.get("calCalorico", ""),
            item.get("horarioInicio", "06:00"),
            "S" if item.get("suspenso") else "N",
            item.get("updatedAt", "")
        ]
        ws_censo.append(row)

    # -------------------------------------------------------------
    # 2. ABA DB_Dietas (Catálogo Oficial de Fórmulas e Diluições)
    # -------------------------------------------------------------
    ws_dietas = wb.create_sheet(title="DB_Dietas")
    headers_dietas = [
        "id", "nome", "categoria", "categoriaNome", "g_po_100ml", "ml_agua_100ml",
        "peso_lata_g", "kcal_100ml", "densidade_padrao", "temperatura_preparo", "instrucoes"
    ]
    ws_dietas.append(headers_dietas)
    
    header_fill_dietas = PatternFill(start_color="166534", end_color="166534", fill_type="solid")
    for col_idx in range(1, len(headers_dietas) + 1):
        cell = ws_dietas.cell(row=1, column=col_idx)
        cell.fill = header_fill_dietas
        cell.font = header_font
        cell.alignment = align_center

    dietas_path = os.path.join(BASE_DIR, "js", "data", "dietas-padrao.js")
    dietas_data = parse_js_array(dietas_path, "DIETAS_PADRAO")
    for d in dietas_data:
        row = [
            d.get("id", ""),
            d.get("nome", ""),
            d.get("categoria", ""),
            d.get("categoriaNome", ""),
            d.get("g_po_100ml", 0),
            d.get("ml_agua_100ml", 0),
            d.get("peso_lata_g", 0),
            d.get("kcal_100ml", 0),
            d.get("densidade_padrao", ""),
            d.get("temperatura_preparo", ""),
            d.get("instrucoes", "")
        ]
        ws_dietas.append(row)

    # -------------------------------------------------------------
    # 3. ABA DB_Enfermarias (47 Enfermarias Oficiais HSP/SPDM)
    # -------------------------------------------------------------
    ws_enf = wb.create_sheet(title="DB_Enfermarias")
    headers_enf = ["id", "nome", "sigla", "leitoInicial", "leitoFinal", "andar"]
    ws_enf.append(headers_enf)
    
    header_fill_enf = PatternFill(start_color="6B21A8", end_color="6B21A8", fill_type="solid")
    for col_idx in range(1, len(headers_enf) + 1):
        cell = ws_enf.cell(row=1, column=col_idx)
        cell.fill = header_fill_enf
        cell.font = header_font
        cell.alignment = align_center

    enf_path = os.path.join(BASE_DIR, "js", "data", "enfermarias-spdm.js")
    enf_data = parse_js_array(enf_path, "ENFERMARIAS_SPDM")
    for item in enf_data:
        row = [
            item.get("id", ""),
            item.get("nome", ""),
            item.get("sigla", ""),
            item.get("leitoInicial", ""),
            item.get("leitoFinal", ""),
            item.get("andar", "")
        ]
        ws_enf.append(row)

    # -------------------------------------------------------------
    # 4. ABA DB_Logs (Auditoria e Histórico de Operações)
    # -------------------------------------------------------------
    ws_logs = wb.create_sheet(title="DB_Logs")
    headers_logs = ["timestamp", "usuario", "operacao", "detalhes"]
    ws_logs.append(headers_logs)
    header_fill_logs = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    for col_idx in range(1, len(headers_logs) + 1):
        cell = ws_logs.cell(row=1, column=col_idx)
        cell.fill = header_fill_logs
        cell.font = header_font
        cell.alignment = align_center

    ws_logs.append(["2026-08-20T08:00:00.000Z", "Sistema HSP", "INICIALIZACAO", "Planilha base criada com sucesso."])

    # Remover a aba vazia padrão
    if default_sheet.title in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(default_sheet)

    # Congelar linha de cabeçalho e aplicar auto-width e bordas
    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = cell_font
                cell.border = thin_border
                if isinstance(cell.value, (int, float)):
                    cell.alignment = align_center

        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 14)

    output_path = os.path.join(BASE_DIR, "Lactario_Digital_Database_HSP.xlsx")
    wb.save(output_path)
    print(f"✅ Planilha criada com sucesso em: {output_path}")
    print(f"   - DB_Censo: {len(censo_data)} pacientes")
    print(f"   - DB_Dietas: {len(dietas_data)} fórmulas/dietas oficiais")
    print(f"   - DB_Enfermarias: {len(enf_data)} enfermarias oficiais")
    print(f"   - DB_Logs: 1 registro de log")

if __name__ == "__main__":
    create_database_spreadsheet()
