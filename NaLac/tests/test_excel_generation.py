#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Test Suite: Validação da Planilha Excel do Lactário (CENSO PARA SPDM, PRODUÇÃO e Turnos)
"""

import os
import sys
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
import openpyxl
from backend.excel_processor import classificar_turno_horario, calcular_distribuicao_preparos

def test_classificacao_turnos():
    """Testa a regra de negócio de turnos de preparo:
       - Preparo 1: 08:00 às 18:00
       - Preparo 2: 20:00 às 06:00 (do dia seguinte)
    """
    assert classificar_turno_horario('08:00') == 1, "08:00 deve ser Preparo 1"
    assert classificar_turno_horario('09:00') == 1, "09:00 deve ser Preparo 1"
    assert classificar_turno_horario('12:00') == 1, "12:00 deve ser Preparo 1"
    assert classificar_turno_horario('15:00') == 1, "15:00 deve ser Preparo 1"
    assert classificar_turno_horario('18:00') == 1, "18:00 deve ser Preparo 1"

    assert classificar_turno_horario('20:00') == 2, "20:00 deve ser Preparo 2"
    assert classificar_turno_horario('21:00') == 2, "21:00 deve ser Preparo 2"
    assert classificar_turno_horario('24:00') == 2, "24:00 deve ser Preparo 2"
    assert classificar_turno_horario('00:00') == 2, "00:00 deve ser Preparo 2"
    assert classificar_turno_horario('03:00') == 2, "03:00 deve ser Preparo 2"
    assert classificar_turno_horario('06:00') == 2, "06:00 deve ser Preparo 2"

    prep1, prep2 = calcular_distribuicao_preparos(['06:00', '09:00', '12:00', '15:00', '18:00', '21:00', '00:00', '03:00'])
    assert prep1 == 4, f"Esperado 4 no preparo 1, obtido {prep1}"
    assert prep2 == 4, f"Esperado 4 no preparo 2, obtido {prep2}"
    print("✅ Teste de Classificação de Turnos: OK!")

def test_estrutura_planilha():
    """Valida a integridade das abas, fórmulas e formatação no arquivo .xlsx."""
    excel_path = 'Copia lactario.xlsx'
    assert os.path.exists(excel_path), f"Arquivo {excel_path} não encontrado!"

    wb = openpyxl.load_workbook(excel_path, data_only=False)
    
    # 1. Validação de Abas Obrigatórias
    abas_esperadas = [
        'CENSO PARA SPDM',
        'PRODUÇÃO',
        'AUTOCLAVADA',
        'NAO AUTOCLAVADA',
        'DIETA ESPECIAL',
        'SOMA AUTOCLAVADA',
        'SOMA NAO AUTOCLAVADA',
        'SOMA ENTERAL',
        'CENSO DISPOS',
        'CENSO VOLUME AUT',
        'CENSO VOLUME N AUT',
        'ABREVIAÇÃO JE',
        'Concentrações e medidas',
        'Config',
        'PEDIDO'
    ]
    for aba in abas_esperadas:
        assert aba in wb.sheetnames, f"Aba obrigatória '{aba}' ausente na planilha!"
    print("✅ Teste de Presença de Todas as Abas: OK!")

    # 2. Validação da Aba 'CENSO PARA SPDM'
    ws_spdm = wb['CENSO PARA SPDM']
    assert "HOSPITAL SÃO PAULO" in str(ws_spdm['A1'].value), "Título principal ausente em CENSO PARA SPDM"
    assert "CENSO CONSOLIDADO PARA SPDM" in str(ws_spdm['A2'].value), "Subtítulo ausente em CENSO PARA SPDM"
    assert ws_spdm.max_row >= 180, f"CENSO PARA SPDM deve conter as 4 seções integradas (linhas={ws_spdm.max_row})"
    print("✅ Teste da Aba 'CENSO PARA SPDM' (4 Seções Integradas): OK!")

    # 3. Validação da Aba 'PRODUÇÃO'
    ws_prod = wb['PRODUÇÃO']
    assert "RELATÓRIO OFICIAL DE PRODUÇÃO" in str(ws_prod['A1'].value), "Título de PRODUÇÃO incorreto"
    assert ws_prod['A5'].value == "Categoria / Dieta", "Cabeçalho de PRODUÇÃO incorreto"
    assert ws_prod['D5'].value == "Preparo 1 (08h-18h) Vol", "Coluna Preparo 1 incorreta em PRODUÇÃO"
    assert ws_prod['E5'].value == "Preparo 2 (20h-06h) Vol", "Coluna Preparo 2 incorreta em PRODUÇÃO"
    
    # Verifica fórmulas de ligação e totais em PRODUÇÃO
    assert str(ws_prod['D6'].value).startswith("="), "Fórmula de Preparo 1 ausente em D6"
    assert str(ws_prod['E6'].value).startswith("="), "Fórmula de Preparo 2 ausente em E6"
    assert "SUM(" in str(ws_prod['F28'].value), "Fórmula de Soma Total ausente no rodapé de PRODUÇÃO"
    print("✅ Teste da Aba 'PRODUÇÃO' (Fórmulas e Totais): OK!")

    # 4. Validação de Fórmulas em AUTOCLAVADA e NÃO AUTOCLAVADA
    ws_auto = wb['AUTOCLAVADA']
    assert ws_auto['L5'].value == "=K5*G5", "Fórmula de Preparo 1 em AUTOCLAVADA!L5 incorreta"
    assert ws_auto['N5'].value == "=M5*G5", "Fórmula de Preparo 2 em AUTOCLAVADA!N5 incorreta"
    assert ws_auto['O5'].value == "=G5*E5", "Fórmula de Volume Total em AUTOCLAVADA!O5 incorreta"
    print("✅ Teste de Fórmulas Dinâmicas em Pacientes: OK!")

    print("\n🎉 TODOS OS TESTES DA ENGENHARIA DE DADOS EM EXCEL PASSARAM COM SUCESSO (100%)!")

if __name__ == '__main__':
    test_classificacao_turnos()
    test_estrutura_planilha()
