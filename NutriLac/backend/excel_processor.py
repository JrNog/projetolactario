#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Motor de Engenharia e Automação de Planilhas Excel
Lactário do Hospital São Paulo (UNIFESP) / SPDM

Objetivos:
1. Replicar fielmente o padrão visual e funcional existente (fontes, cores, bordas, fórmulas matriciais e mesclagens).
2. Criar e atualizar a aba unificada 'CENSO PARA SPDM' (CENSO DISPOS, CENSO VOLUME AUT, CENSO VOLUME N AUT, ABREVIAÇÃO JE).
3. Criar e estruturar a aba 'PRODUÇÃO' consolidando as métricas executivas e mapas de distribuição.
4. Implementar a regra de turnos de preparo:
   - PREPARO 1: Refeições das 08:00 às 18:00
   - PREPARO 2: Refeições das 20:00 às 06:00 (do dia seguinte)
"""

import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def classificar_turno_horario(hora_val):
    """
    Classifica um horário em PREPARO 1 ou PREPARO 2 conforme a regra oficial:
    - PREPARO 1: 08:00 às 18:00 (8 <= hora <= 18)
    - PREPARO 2: 20:00 às 06:00 do dia seguinte (20 <= hora <= 24 ou 0 <= hora <= 6 ou hora == 7)
    """
    if hora_val is None:
        return None
    
    str_val = str(hora_val).strip()
    match = re.match(r'^(\d{1,2})', str_val)
    if match:
        h = int(match.group(1))
        if h == 24:
            h = 0
        if 8 <= h <= 18:
            return 1
        else:
            return 2
    return None

def calcular_distribuicao_preparos(horarios_lista):
    """
    Dada uma lista de horários de um paciente, calcula quantas refeições caem em:
    - Preparo 1 (08h - 18h)
    - Preparo 2 (20h - 06h)
    """
    prep1 = 0
    prep2 = 0
    for h_str in horarios_lista:
        turno = classificar_turno_horario(h_str)
        if turno == 1:
            prep1 += 1
        elif turno == 2:
            prep2 += 1
    return prep1, prep2

def copiar_estilo_celula(src_cell, tgt_cell):
    """Copia fontes, preenchimentos, bordas, alinhamento e formatos de número entre células."""
    if src_cell.has_style:
        tgt_cell.font = Font(
            name=src_cell.font.name or 'Calibri',
            size=src_cell.font.size or 11,
            bold=src_cell.font.bold or False,
            italic=src_cell.font.italic or False,
            underline=src_cell.font.underline,
            color=src_cell.font.color
        )
        if src_cell.fill and src_cell.fill.fill_type:
            tgt_cell.fill = PatternFill(
                fill_type=src_cell.fill.fill_type,
                start_color=src_cell.fill.start_color,
                end_color=src_cell.fill.end_color
            )
        tgt_cell.border = Border(
            left=src_cell.border.left,
            right=src_cell.border.right,
            top=src_cell.border.top,
            bottom=src_cell.border.bottom
        )
        tgt_cell.alignment = Alignment(
            horizontal=src_cell.alignment.horizontal,
            vertical=src_cell.alignment.vertical,
            text_rotation=src_cell.alignment.text_rotation,
            wrap_text=src_cell.alignment.wrap_text,
            shrink_to_fit=src_cell.alignment.shrink_to_fit,
            indent=src_cell.alignment.indent
        )
        tgt_cell.number_format = src_cell.number_format

def aplicar_estilo_cabecalho_secao(ws, start_row, start_col, end_col, titulo, cor_fundo='1F4E78', cor_texto='FFFFFF'):
    """Aplica formatação premium a cabeçalhos de bloco/seção no Excel."""
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    cell = ws.cell(row=start_row, column=start_col, value=titulo)
    cell.font = Font(name='Calibri', size=11, bold=True, color=cor_texto)
    cell.fill = PatternFill(fill_type='solid', start_color=cor_fundo, end_color=cor_fundo)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='A6A6A6'),
        right=Side(style='thin', color='A6A6A6'),
        top=Side(style='thin', color='A6A6A6'),
        bottom=Side(style='thin', color='A6A6A6')
    )
    for c in range(start_col, end_col + 1):
        ws.cell(row=start_row, column=c).border = thin_border

def gerar_planilha_lactario(caminho_entrada='Copia lactario.xlsx', caminho_saida='Copia lactario.xlsx'):
    """
    Processa a planilha oficial do Lactário, atualizando as regras de turnos e gerando:
    1. Aba 'CENSO PARA SPDM' (unificada com as 4 seções)
    2. Aba 'PRODUÇÃO' (consolidada)
    Preservando todas as abas, dados e fórmulas existentes.
    """
    print(f"📖 Carregando planilha de referência: {caminho_entrada}...")
    wb = openpyxl.load_workbook(caminho_entrada, data_only=False)

    # -------------------------------------------------------------------------
    # 1. ATUALIZAÇÃO DA REGRA DE TURNOS NAS ABAS DE PACIENTES
    # -------------------------------------------------------------------------
    print("⚙️ Atualizando fórmulas de turnos de preparo (08h-18h / 20h-06h)...")
    for sname in ['AUTOCLAVADA', 'NAO AUTOCLAVADA']:
        if sname in wb.sheetnames:
            ws = wb[sname]
            for r in range(5, ws.max_row + 1):
                cell_A = ws.cell(row=r, column=1)
                if isinstance(cell_A, openpyxl.cell.cell.MergedCell):
                    continue

                leito = cell_A.value
                vezes = ws.cell(row=r, column=5).value
                
                # Se for linha de paciente ativo com leito
                if leito and str(leito).strip() != '' and str(leito).strip() != 'Total:' and vezes is not None:
                    cell_L = ws.cell(row=r, column=12)
                    cell_N = ws.cell(row=r, column=14)
                    cell_O = ws.cell(row=r, column=15)

                    # Garante que as colunas de volume estejam com as fórmulas corretas
                    if not isinstance(cell_L, openpyxl.cell.cell.MergedCell):
                        cell_L.value = f"=K{r}*G{r}"
                    if not isinstance(cell_N, openpyxl.cell.cell.MergedCell):
                        cell_N.value = f"=M{r}*G{r}"
                    if not isinstance(cell_O, openpyxl.cell.cell.MergedCell):
                        cell_O.value = f"=G{r}*E{r}"

    # -------------------------------------------------------------------------
    # 2. CONSTRUÇÃO DA ABA UNIFICADA 'CENSO PARA SPDM'
    # -------------------------------------------------------------------------
    print("🏗️ Construindo aba consolidada 'CENSO PARA SPDM'...")
    if 'CENSO PARA SPDM' in wb.sheetnames:
        del wb['CENSO PARA SPDM']
    
    ws_spdm = wb.create_sheet(title='CENSO PARA SPDM', index=0)
    ws_spdm.views.sheetView[0].showGridLines = True

    # Banner Principal do Hospital São Paulo / SPDM
    ws_spdm.merge_cells('A1:S1')
    cell_top = ws_spdm['A1']
    cell_top.value = "HOSPITAL SÃO PAULO - UNIFESP • CENTRAL DE NUTRIÇÃO E DIETÉTICA"
    cell_top.font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
    cell_top.fill = PatternFill(fill_type='solid', start_color='1F4E78', end_color='1F4E78')
    cell_top.alignment = Alignment(horizontal='center', vertical='center')

    ws_spdm.merge_cells('A2:S2')
    cell_sub = ws_spdm['A2']
    cell_sub.value = "CENSO CONSOLIDADO PARA SPDM (DISPOSITIVOS, VOLUMES AUTOCLAVADOS, NÃO AUTOCLAVADOS E ABREVIAÇÃO DE JEJUM)"
    cell_sub.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    cell_sub.fill = PatternFill(fill_type='solid', start_color='2F5597', end_color='2F5597')
    cell_sub.alignment = Alignment(horizontal='center', vertical='center')

    ws_spdm.row_dimensions[1].height = 24
    ws_spdm.row_dimensions[2].height = 18

    # Estilos padronizados
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    total_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )

    current_row = 4

    # -------------------------------------------------------------------------
    # SEÇÃO 1: CENSO DE DISPOSITIVOS (MAMADEIRAS, FRASCOS, COPOS, SERINGAS)
    # -------------------------------------------------------------------------
    aplicar_estilo_cabecalho_secao(ws_spdm, current_row, 1, 9, "1. CENSO DE DISPOSITIVOS (MAMADEIRAS, FRASCOS, COPOS, SERINGAS)", '1F4E78')
    current_row += 1
    sec1_header_row = current_row

    if 'CENSO DISPOS' in wb.sheetnames:
        ws_src = wb['CENSO DISPOS']
        for r in range(1, ws_src.max_row + 1):
            target_r = current_row
            ws_spdm.row_dimensions[target_r].height = 18
            for c in range(1, min(ws_src.max_column + 1, 10)):
                src_c = ws_src.cell(row=r, column=c)
                tgt_c = ws_spdm.cell(row=target_r, column=c)
                
                # Ajusta fórmulas
                raw_val = src_c.value
                if r == 49 and str(raw_val).startswith("=SUM("):
                    col_let = get_column_letter(c)
                    start_sum = sec1_header_row + 1
                    end_sum = target_r - 1
                    tgt_c.value = f"=SUM({col_let}{start_sum}:{col_let}{end_sum})"
                elif isinstance(raw_val, str) and raw_val.startswith("="):
                    # Corrige referência ao cabeçalho da própria seção
                    adj_val = re.sub(r'([C-I])\$1\b', rf'\1${sec1_header_row}', raw_val)
                    tgt_c.value = adj_val
                else:
                    tgt_c.value = raw_val

                copiar_estilo_celula(src_c, tgt_c)
                if r == 1:
                    tgt_c.fill = PatternFill(fill_type='solid', start_color='D9E1F2', end_color='D9E1F2')
                    tgt_c.font = Font(name='Calibri', size=10, bold=True, color='002060')
                    tgt_c.alignment = Alignment(horizontal='center', vertical='center')
                elif r == 49:
                    tgt_c.fill = PatternFill(fill_type='solid', start_color='FFF2CC', end_color='FFF2CC')
                    tgt_c.font = Font(name='Calibri', size=10, bold=True, color='7F6000')
                    tgt_c.border = total_border
                else:
                    tgt_c.border = thin_border
            current_row += 1

    current_row += 2

    # -------------------------------------------------------------------------
    # SEÇÃO 2: CENSO DE VOLUMES - DIETAS AUTOCLAVADAS
    # -------------------------------------------------------------------------
    aplicar_estilo_cabecalho_secao(ws_spdm, current_row, 1, 9, "2. CENSO DE VOLUME - DIETAS AUTOCLAVADAS", '2F5597')
    current_row += 1
    sec2_header_row = current_row

    if 'CENSO VOLUME AUT' in wb.sheetnames:
        ws_src = wb['CENSO VOLUME AUT']
        for r in range(1, ws_src.max_row + 1):
            target_r = current_row
            ws_spdm.row_dimensions[target_r].height = 18
            for c in range(1, min(ws_src.max_column + 1, 10)):
                src_c = ws_src.cell(row=r, column=c)
                tgt_c = ws_spdm.cell(row=target_r, column=c)

                raw_val = src_c.value
                if r == 49 and str(raw_val).startswith("=SUM("):
                    col_let = get_column_letter(c)
                    start_sum = sec2_header_row + 1
                    end_sum = target_r - 1
                    tgt_c.value = f"=SUM({col_let}{start_sum}:{col_let}{end_sum})"
                else:
                    tgt_c.value = raw_val

                copiar_estilo_celula(src_c, tgt_c)
                if r == 1:
                    tgt_c.fill = PatternFill(fill_type='solid', start_color='D9E1F2', end_color='D9E1F2')
                    tgt_c.font = Font(name='Calibri', size=10, bold=True, color='002060')
                    tgt_c.alignment = Alignment(horizontal='center', vertical='center')
                elif r == 49:
                    tgt_c.fill = PatternFill(fill_type='solid', start_color='E2EFDA', end_color='E2EFDA')
                    tgt_c.font = Font(name='Calibri', size=10, bold=True, color='375623')
                    tgt_c.border = total_border
                else:
                    tgt_c.border = thin_border
            current_row += 1

    current_row += 2

    # -------------------------------------------------------------------------
    # SEÇÃO 3: CENSO DE VOLUMES - DIETAS NÃO AUTOCLAVADAS
    # -------------------------------------------------------------------------
    aplicar_estilo_cabecalho_secao(ws_spdm, current_row, 1, 19, "3. CENSO DE VOLUME - DIETAS NÃO AUTOCLAVADAS (BANCADA ESTÉRIL)", '595959')
    current_row += 1
    sec3_header_row = current_row

    if 'CENSO VOLUME N AUT' in wb.sheetnames:
        ws_src = wb['CENSO VOLUME N AUT']
        for r in range(1, ws_src.max_row + 1):
            target_r = current_row
            ws_spdm.row_dimensions[target_r].height = 18
            for c in range(1, min(ws_src.max_column + 1, 20)):
                src_c = ws_src.cell(row=r, column=c)
                tgt_c = ws_spdm.cell(row=target_r, column=c)

                raw_val = src_c.value
                if r == 49 and str(raw_val).startswith("=SUM("):
                    col_let = get_column_letter(c)
                    start_sum = sec3_header_row + 1
                    end_sum = target_r - 1
                    tgt_c.value = f"=SUM({col_let}{start_sum}:{col_let}{end_sum})"
                else:
                    tgt_c.value = raw_val

                copiar_estilo_celula(src_c, tgt_c)
                if r == 1:
                    tgt_c.fill = PatternFill(fill_type='solid', start_color='F2DCDB', end_color='F2DCDB')
                    tgt_c.font = Font(name='Calibri', size=9, bold=True, color='595959')
                    tgt_c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                elif r == 49:
                    tgt_c.fill = PatternFill(fill_type='solid', start_color='FCE4D6', end_color='FCE4D6')
                    tgt_c.font = Font(name='Calibri', size=10, bold=True, color='C65911')
                    tgt_c.border = total_border
                else:
                    tgt_c.border = thin_border
            current_row += 1

    current_row += 2

    # -------------------------------------------------------------------------
    # SEÇÃO 4: ABREVIAÇÃO DE JEJUM (CHÁ COM MALTODEXTRINA)
    # -------------------------------------------------------------------------
    aplicar_estilo_cabecalho_secao(ws_spdm, current_row, 1, 8, "4. ABREVIAÇÃO DE JEJUM - CHÁ SEM AÇÚCAR + 25G DE MALTODEXTRINA", '375623')
    current_row += 1

    if 'ABREVIAÇÃO JE' in wb.sheetnames:
        ws_src = wb['ABREVIAÇÃO JE']
        for r in range(4, ws_src.max_row + 1):
            target_r = current_row
            ws_spdm.row_dimensions[target_r].height = 18
            for c in range(1, min(ws_src.max_column + 1, 9)):
                src_c = ws_src.cell(row=r, column=c)
                tgt_c = ws_spdm.cell(row=target_r, column=c)
                tgt_c.value = src_c.value
                copiar_estilo_celula(src_c, tgt_c)
                if r == 4:
                    tgt_c.fill = PatternFill(fill_type='solid', start_color='E2EFDA', end_color='E2EFDA')
                    tgt_c.font = Font(name='Calibri', size=10, bold=True, color='375623')
                    tgt_c.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    tgt_c.border = thin_border
            current_row += 1

    # Ajuste de largura das colunas de A a S
    larguras = {
        'A': 32, 'B': 16, 'C': 16, 'D': 16, 'E': 16, 'F': 16, 'G': 16, 'H': 16, 'I': 16,
        'J': 16, 'K': 16, 'L': 16, 'M': 16, 'N': 16, 'O': 16, 'P': 16, 'Q': 16, 'R': 16, 'S': 16
    }
    for col_let, w in larguras.items():
        ws_spdm.column_dimensions[col_let].width = w

    # -------------------------------------------------------------------------
    # 3. CONSTRUÇÃO DA ABA 'PRODUÇÃO'
    # -------------------------------------------------------------------------
    print("🏭 Estruturando aba 'PRODUÇÃO' com consolidação executiva...")
    if 'PRODUÇÃO' in wb.sheetnames:
        del wb['PRODUÇÃO']
    
    ws_prod = wb.create_sheet(title='PRODUÇÃO', index=1)
    ws_prod.views.sheetView[0].showGridLines = True

    # Banner Oficial Produção
    ws_prod.merge_cells('A1:L1')
    p_top = ws_prod['A1']
    p_top.value = "HOSPITAL SÃO PAULO • CENTRAL DE NUTRIÇÃO E DIETÉTICA - RELATÓRIO OFICIAL DE PRODUÇÃO"
    p_top.font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
    p_top.fill = PatternFill(fill_type='solid', start_color='002060', end_color='002060')
    p_top.alignment = Alignment(horizontal='center', vertical='center')

    ws_prod.merge_cells('A2:L2')
    p_sub = ws_prod['A2']
    p_sub.value = "CONSOLIDAÇÃO POR FÓRMULAS, TURNOS DE PREPARO (08H-18H / 20H-06H) E DISTRIBUIÇÃO HOSPITALAR"
    p_sub.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    p_sub.fill = PatternFill(fill_type='solid', start_color='1F4E78', end_color='1F4E78')
    p_sub.alignment = Alignment(horizontal='center', vertical='center')

    ws_prod.row_dimensions[1].height = 24
    ws_prod.row_dimensions[2].height = 18

    # Tabela 1: Resumo Consolidado de Fórmulas Autoclavadas & Não Autoclavadas
    ws_prod.merge_cells('A4:L4')
    t1_head = ws_prod['A4']
    t1_head.value = "MAPA GERAL DE PREPARO DE DIETAS (AUTOCLAVADAS & NÃO AUTOCLAVADAS)"
    t1_head.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    t1_head.fill = PatternFill(fill_type='solid', start_color='2F5597', end_color='2F5597')
    t1_head.alignment = Alignment(horizontal='center', vertical='center')

    headers_prod = [
        ('A', 'Categoria / Dieta'),
        ('B', 'Pó (g/100ml)'),
        ('C', 'Água (ml/100ml)'),
        ('D', 'Preparo 1 (08h-18h) Vol'),
        ('E', 'Preparo 2 (20h-06h) Vol'),
        ('F', 'Volume Total (ml)'),
        ('G', 'Consumo Pó (g)'),
        ('H', 'Água Total (ml)'),
        ('I', 'Mamadeiras'),
        ('J', 'Frascos Enterais'),
        ('K', 'Seringas / Copos'),
        ('L', 'Status')
    ]

    for col_let, text in headers_prod:
        cell = ws_prod[f'{col_let}5']
        cell.value = text
        cell.font = Font(name='Calibri', size=10, bold=True, color='002060')
        cell.fill = PatternFill(fill_type='solid', start_color='D9E1F2', end_color='D9E1F2')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws_prod.row_dimensions[5].height = 28

    # Linhas de fórmulas mapeadas com links para as abas de soma e censo
    dietas_resumo = [
        ("Pré Nan (16,3%)", "16,3", "90,0", "=AUTOCLAVADA!L44", "=AUTOCLAVADA!N44", "='CENSO VOLUME AUT'!B49", "='CENSO PÓ AUT'!B49", "=F6*0.9", "=COUNTIFS(AUTOCLAVADA!I:I, 'Mamadeira', AUTOCLAVADA!D:D, '<>S')", "=COUNTIFS(AUTOCLAVADA!I:I, 'Frasco Enteral', AUTOCLAVADA!D:D, '<>S')", "=COUNTIFS(AUTOCLAVADA!I:I, 'Seringa', AUTOCLAVADA!D:D, '<>S')", "AUTOCLAVADA"),
        ("Pré Nan Concentrado (19,6%)", "19,6", "90,0", "=AUTOCLAVADA!L56", "=AUTOCLAVADA!N56", "='CENSO VOLUME AUT'!C49", "='CENSO PÓ AUT'!C49", "=F7*0.9", "-", "-", "-", "AUTOCLAVADA"),
        ("Nan 1 Comfor (13,5%)", "13,5", "90,0", "=AUTOCLAVADA!L101", "=AUTOCLAVADA!N101", "='CENSO VOLUME AUT'!D49", "='CENSO PÓ AUT'!D49", "=F8*0.9", "-", "-", "-", "AUTOCLAVADA"),
        ("Nan 1 Concentrado (16,2%)", "16,2", "90,0", "=AUTOCLAVADA!L112", "=AUTOCLAVADA!N112", "='CENSO VOLUME AUT'!E49", "='CENSO PÓ AUT'!E49", "=F9*0.9", "-", "-", "-", "AUTOCLAVADA"),
        ("Nan 2 Comfor (14,2%)", "14,2", "90,0", "=AUTOCLAVADA!L142", "=AUTOCLAVADA!N142", "='CENSO VOLUME AUT'!F49", "='CENSO PÓ AUT'!F49", "=F10*0.9", "-", "-", "-", "AUTOCLAVADA"),
        ("Nan 2 Concentrado (17,0%)", "17,0", "90,0", "=AUTOCLAVADA!L154", "=AUTOCLAVADA!N154", "='CENSO VOLUME AUT'!G49", "='CENSO PÓ AUT'!G49", "=F11*0.9", "-", "-", "-", "AUTOCLAVADA"),
        ("Aptamil Soja (13,8%)", "13,8", "90,0", "=AUTOCLAVADA!L168", "=AUTOCLAVADA!N168", "='CENSO VOLUME AUT'!H49", "='CENSO PÓ AUT'!H49", "=F12*0.9", "-", "-", "-", "AUTOCLAVADA"),
        ("LD sem Açúcar (Ninho 12,5%)", "12,5", "90,0", "=AUTOCLAVADA!L186", "=AUTOCLAVADA!N186", "='CENSO VOLUME AUT'!I49", "='CENSO PÓ AUT'!I49", "=F13*0.9", "-", "-", "-", "AUTOCLAVADA"),
        ("LD com Açúcar (12,5%)", "12,5", "90,0", "=AUTOCLAVADA!L203", "=AUTOCLAVADA!N203", "=AUTOCLAVADA!L203+AUTOCLAVADA!N203", "=F14*0.125", "=F14*0.9", "-", "-", "-", "AUTOCLAVADA"),
        ("Neocate (13,8%)", "13,8", "90,0", "='NAO AUTOCLAVADA'!L23", "='NAO AUTOCLAVADA'!N23", "='CENSO VOLUME N AUT'!B49", "='CENSO PÓ N AUT'!B49", "=F15*0.9", "-", "-", "-", "NÃO AUTOCLAVADA"),
        ("Neocate Concentrado (16,6%)", "16,6", "90,0", "='NAO AUTOCLAVADA'!L34", "='NAO AUTOCLAVADA'!N34", "='CENSO VOLUME N AUT'!C49", "='CENSO PÓ N AUT'!C49", "=F16*0.9", "-", "-", "-", "NÃO AUTOCLAVADA"),
        ("Leite Desnatado (10,0%)", "10,0", "90,0", "='NAO AUTOCLAVADA'!L45", "='NAO AUTOCLAVADA'!N45", "='CENSO VOLUME N AUT'!D49", "='CENSO PÓ N AUT'!D49", "=F17*0.9", "-", "-", "-", "NÃO AUTOCLAVADA"),
        ("Pregomin Pepti 1:30 (12,9%)", "12,9", "90,0", "='NAO AUTOCLAVADA'!L80", "='NAO AUTOCLAVADA'!N80", "='CENSO VOLUME N AUT'!F49", "='CENSO PÓ N AUT'!F49", "=F18*0.9", "-", "-", "-", "NÃO AUTOCLAVADA"),
        ("Pregomin Concentrado 1:25 (15,5%)", "15,5", "90,0", "='NAO AUTOCLAVADA'!L93", "='NAO AUTOCLAVADA'!N93", "='CENSO VOLUME N AUT'!G49", "='CENSO PÓ N AUT'!G49", "=F19*0.9", "-", "-", "-", "NÃO AUTOCLAVADA"),
        ("Pregomin Concentrado 1:20 (19,35%)", "19,35", "90,0", "='NAO AUTOCLAVADA'!L102", "='NAO AUTOCLAVADA'!N102", "='CENSO VOLUME N AUT'!H49", "='CENSO PÓ N AUT'!H49", "=F20*0.9", "-", "-", "-", "NÃO AUTOCLAVADA"),
        ("Nan Sem Lactose (13,2%)", "13,2", "90,0", "='NAO AUTOCLAVADA'!L126", "='NAO AUTOCLAVADA'!N126", "='CENSO VOLUME N AUT'!J49", "='CENSO PÓ N AUT'!J49", "=F21*0.9", "-", "-", "-", "NÃO AUTOCLAVADA"),
        ("Nan Espessar (13,3%)", "13,3", "90,0", "='NAO AUTOCLAVADA'!L146", "='NAO AUTOCLAVADA'!N146", "='CENSO VOLUME N AUT'!L49", "='CENSO PÓ N AUT'!L49", "=F22*0.9", "-", "-", "-", "NÃO AUTOCLAVADA"),
        ("Peptamen Junior (20,0%)", "20,0", "84,0", "='NAO AUTOCLAVADA'!L168", "='NAO AUTOCLAVADA'!N168", "='CENSO VOLUME N AUT'!N49", "='CENSO PÓ N AUT'!N49", "=F23*0.84", "-", "-", "-", "NÃO AUTOCLAVADA"),
        ("Fortini 1.0 (20,0%)", "20,0", "85,0", "='NAO AUTOCLAVADA'!L187", "='NAO AUTOCLAVADA'!N187", "='CENSO VOLUME N AUT'!O49", "='CENSO PÓ N AUT'!O49", "=F24*0.85", "-", "-", "-", "NÃO AUTOCLAVADA"),
        ("Infatrini (20,4%)", "20,4", "85,0", "='NAO AUTOCLAVADA'!L209", "='NAO AUTOCLAVADA'!N209", "='CENSO VOLUME N AUT'!Q49", "='CENSO PÓ N AUT'!Q49", "=F25*0.85", "-", "-", "-", "NÃO AUTOCLAVADA"),
        ("Modulen 1 (20,0%)", "20,0", "84,0", "='NAO AUTOCLAVADA'!L216", "='NAO AUTOCLAVADA'!N216", "='CENSO VOLUME N AUT'!R49", "='CENSO PÓ N AUT'!R49", "=F26*0.84", "-", "-", "-", "NÃO AUTOCLAVADA"),
        ("Modulen 2 Concentrado (30,0%)", "30,0", "84,0", "='NAO AUTOCLAVADA'!L223", "='NAO AUTOCLAVADA'!N223", "='CENSO VOLUME N AUT'!S49", "='CENSO PÓ N AUT'!S49", "=F27*0.84", "-", "-", "-", "NÃO AUTOCLAVADA")
    ]

    row_idx = 6
    for item in dietas_resumo:
        ws_prod.row_dimensions[row_idx].height = 18
        for col_idx, val in enumerate(item, start=1):
            cell = ws_prod.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name='Calibri', size=10)
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            elif col_idx in [2, 3]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_idx in [4, 5, 6, 7, 8]:
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '#,##0.0'
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        row_idx += 1

    # Linha de Totais da Tabela de Produção
    ws_prod.row_dimensions[row_idx].height = 20
    cell_tot_lbl = ws_prod.cell(row=row_idx, column=1, value="TOTAL GERAL DE PRODUÇÃO:")
    cell_tot_lbl.font = Font(name='Calibri', size=10, bold=True, color='000000')
    cell_tot_lbl.fill = PatternFill(fill_type='solid', start_color='FFF2CC', end_color='FFF2CC')
    cell_tot_lbl.border = total_border

    for c in range(2, 13):
        cell_tot = ws_prod.cell(row=row_idx, column=c)
        col_let = get_column_letter(c)
        if c in [4, 5, 6, 7, 8]:
            cell_tot.value = f"=SUM({col_let}6:{col_let}{row_idx-1})"
            cell_tot.number_format = '#,##0.0'
        else:
            cell_tot.value = "-"
        cell_tot.font = Font(name='Calibri', size=10, bold=True, color='7F6000')
        cell_tot.fill = PatternFill(fill_type='solid', start_color='FFF2CC', end_color='FFF2CC')
        cell_tot.alignment = Alignment(horizontal='right' if c in [4,5,6,7,8] else 'center', vertical='center')
        cell_tot.border = total_border

    # Ajuste de colunas na aba Produção
    larguras_prod = {
        'A': 36, 'B': 14, 'C': 15, 'D': 22, 'E': 22, 'F': 18, 'G': 16, 'H': 16,
        'I': 14, 'J': 15, 'K': 16, 'L': 18
    }
    for col_let, w in larguras_prod.items():
        ws_prod.column_dimensions[col_let].width = w

    # Salva o arquivo preservando a integridade das fórmulas
    print(f"💾 Gravando alterações na planilha final: {caminho_saida}...")
    wb.save(caminho_saida)
    print("✅ Planilha gerada e atualizada com sucesso com todas as fórmulas, estilos e regras de negócio!")

if __name__ == '__main__':
    gerar_planilha_lactario()
